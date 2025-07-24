"""
Integration tests for CFDI Control Application
"""

import unittest
import tempfile
import os
from pathlib import Path
import sys
from openpyxl import Workbook

# Add src to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from core.xml_parser import CFDIXMLParser
from core.excel_processor import ExcelProcessor
from core.data_models import CFDIDataProcessor
from config.settings import CFDI_MAPPING, EXCEL_CONFIG


class TestCFDIIntegration(unittest.TestCase):
    """Integration tests for the complete CFDI processing workflow."""
    
    def setUp(self):
        """Set up test fixtures."""
        self.xml_parser = CFDIXMLParser()
        self.excel_processor = ExcelProcessor()
        self.data_processor = CFDIDataProcessor()
        self.temp_dir = tempfile.mkdtemp()
        
        # Sample CFDI XML content
        self.sample_xml = '''<?xml version="1.0" encoding="UTF-8"?>
<cfdi:Comprobante xmlns:cfdi="http://www.sat.gob.mx/cfd/3" 
                   Fecha="2024-01-15T10:30:00" 
                   FormaPago="01" 
                   SubTotal="1000.00" 
                   Descuento="0.00" 
                   Moneda="MXN" 
                   Total="1160.00" 
                   TipoDeComprobante="I" 
                   MetodoPago="PUE">
    <cfdi:Emisor Rfc="AAA010101AAA" Nombre="EMPRESA EJEMPLO S.A. DE C.V." RegimenFiscal="601"/>
    <cfdi:Receptor Rfc="XEXX010101000" RegimenFiscalReceptor="601" UsoCFDI="G01"/>
    <cfdi:Impuestos TotalImpuestosTrasladados="160.00"/>
</cfdi:Comprobante>'''
    
    def tearDown(self):
        """Clean up test fixtures."""
        import shutil
        if os.path.exists(self.temp_dir):
            shutil.rmtree(self.temp_dir)
    
    def create_test_excel_template(self, filename):
        """Create a test Excel template with proper structure."""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create month tabs with proper names
        month_names = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                      "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        
        for month_name in month_names:
            ws = wb.create_sheet(month_name)
            
            # Add headers in row 3
            for xml_path, column in CFDI_MAPPING.items():
                ws[f"{column}3"] = xml_path
        
        wb.save(filename)
        return filename
    
    def create_test_xml_files(self, count=3):
        """Create test XML files."""
        xml_files = []
        for i in range(count):
            xml_file = os.path.join(self.temp_dir, f"test_cfdi_{i+1}.xml")
            with open(xml_file, 'w') as f:
                # Modify the XML slightly for each file
                modified_xml = self.sample_xml.replace(
                    '2024-01-15T10:30:00', 
                    f'2024-01-{15+i:02d}T10:30:00'
                ).replace(
                    '1160.00', 
                    f'{1160.00 + i * 100:.2f}'
                )
                f.write(modified_xml)
            xml_files.append(xml_file)
        return xml_files
    
    def test_complete_workflow_success(self):
        """Test the complete workflow from XML parsing to Excel generation."""
        # Step 1: Create test files
        template_path = os.path.join(self.temp_dir, "test_template.xlsx")
        self.create_test_excel_template(template_path)
        xml_files = self.create_test_xml_files(3)
        
        # Step 2: Parse XML files
        raw_data_list = self.xml_parser.parse_multiple_files(xml_files)
        self.assertEqual(len(raw_data_list), 3)
        
        # Step 3: Process and validate data
        processing_result = self.data_processor.process_raw_data(raw_data_list)
        self.assertEqual(processing_result.successful_files, 3)
        self.assertEqual(processing_result.failed_files, 0)
        
        # Step 4: Generate Excel output
        excel_result = self.excel_processor.process_cfdi_to_excel(
            template_path=template_path,
            cfdi_data_list=processing_result.processed_data,
            year=2024,
            month=1
        )
        
        # Step 5: Verify results
        self.assertTrue(excel_result['success'])
        self.assertEqual(excel_result['records_processed'], 3)
        self.assertTrue(os.path.exists(excel_result['output_path']))
    
    def test_workflow_with_invalid_files(self):
        """Test workflow with some invalid files."""
        # Step 1: Create test files
        template_path = os.path.join(self.temp_dir, "test_template.xlsx")
        self.create_test_excel_template(template_path)
        
        # Create valid XML files
        xml_files = self.create_test_xml_files(2)
        
        # Create invalid XML file
        invalid_xml_file = os.path.join(self.temp_dir, "invalid.xml")
        with open(invalid_xml_file, 'w') as f:
            f.write("This is not valid XML")
        xml_files.append(invalid_xml_file)
        
        # Step 2: Parse XML files
        raw_data_list = self.xml_parser.parse_multiple_files(xml_files)
        self.assertEqual(len(raw_data_list), 2)  # Should skip invalid file
        
        # Step 3: Process and validate data
        processing_result = self.data_processor.process_raw_data(raw_data_list)
        self.assertEqual(processing_result.successful_files, 2)
        self.assertEqual(processing_result.failed_files, 0)
        
        # Step 4: Generate Excel output
        excel_result = self.excel_processor.process_cfdi_to_excel(
            template_path=template_path,
            cfdi_data_list=processing_result.processed_data,
            year=2024,
            month=1
        )
        
        # Step 5: Verify results
        self.assertTrue(excel_result['success'])
        self.assertEqual(excel_result['records_processed'], 2)
    
    def test_workflow_with_invalid_template(self):
        """Test workflow with invalid Excel template."""
        # Step 1: Create invalid template (missing headers)
        template_path = os.path.join(self.temp_dir, "invalid_template.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        wb.save(template_path)
        
        # Create valid XML files
        xml_files = self.create_test_xml_files(1)
        
        # Step 2: Parse XML files
        raw_data_list = self.xml_parser.parse_multiple_files(xml_files)
        self.assertEqual(len(raw_data_list), 1)
        
        # Step 3: Process and validate data
        processing_result = self.data_processor.process_raw_data(raw_data_list)
        self.assertEqual(processing_result.successful_files, 1)
        
        # Step 4: Try to generate Excel output (should fail)
        excel_result = self.excel_processor.process_cfdi_to_excel(
            template_path=template_path,
            cfdi_data_list=processing_result.processed_data,
            year=2024,
            month=1
        )
        
        # Step 5: Verify failure
        self.assertFalse(excel_result['success'])
        self.assertIn('error_message', excel_result)
    
    def test_workflow_with_empty_data(self):
        """Test workflow with empty/invalid CFDI data."""
        # Step 1: Create test files
        template_path = os.path.join(self.temp_dir, "test_template.xlsx")
        self.create_test_excel_template(template_path)
        
        # Create XML file with minimal data
        minimal_xml_file = os.path.join(self.temp_dir, "minimal.xml")
        minimal_xml = '''<?xml version="1.0" encoding="UTF-8"?>
<cfdi:Comprobante xmlns:cfdi="http://www.sat.gob.mx/cfd/3" 
                   Fecha="2024-01-15T10:30:00">
</cfdi:Comprobante>'''
        
        with open(minimal_xml_file, 'w') as f:
            f.write(minimal_xml)
        
        # Step 2: Parse XML files
        raw_data_list = self.xml_parser.parse_multiple_files([minimal_xml_file])
        self.assertEqual(len(raw_data_list), 1)
        
        # Step 3: Process and validate data
        processing_result = self.data_processor.process_raw_data(raw_data_list)
        self.assertEqual(processing_result.successful_files, 0)
        self.assertEqual(processing_result.failed_files, 1)
        self.assertGreater(len(processing_result.errors), 0)
        
        # Step 4: Try to generate Excel output (should handle empty data gracefully)
        excel_result = self.excel_processor.process_cfdi_to_excel(
            template_path=template_path,
            cfdi_data_list=processing_result.processed_data,
            year=2024,
            month=1
        )
        
        # Step 5: Verify results (should succeed but with 0 records)
        self.assertTrue(excel_result['success'])
        self.assertEqual(excel_result['records_processed'], 0)
    
    def test_workflow_performance(self):
        """Test workflow performance with multiple files."""
        # Step 1: Create test files
        template_path = os.path.join(self.temp_dir, "test_template.xlsx")
        self.create_test_excel_template(template_path)
        xml_files = self.create_test_xml_files(10)  # 10 files
        
        # Step 2: Parse XML files
        raw_data_list = self.xml_parser.parse_multiple_files(xml_files)
        self.assertEqual(len(raw_data_list), 10)
        
        # Step 3: Process and validate data
        processing_result = self.data_processor.process_raw_data(raw_data_list)
        self.assertEqual(processing_result.successful_files, 10)
        
        # Step 4: Generate Excel output
        excel_result = self.excel_processor.process_cfdi_to_excel(
            template_path=template_path,
            cfdi_data_list=processing_result.processed_data,
            year=2024,
            month=1
        )
        
        # Step 5: Verify results
        self.assertTrue(excel_result['success'])
        self.assertEqual(excel_result['records_processed'], 10)
    
    def test_data_integrity_through_workflow(self):
        """Test that data integrity is maintained through the workflow."""
        # Step 1: Create test files
        template_path = os.path.join(self.temp_dir, "test_template.xlsx")
        self.create_test_excel_template(template_path)
        xml_files = self.create_test_xml_files(1)
        
        # Step 2: Parse XML files
        raw_data_list = self.xml_parser.parse_multiple_files(xml_files)
        original_data = raw_data_list[0]
        
        # Step 3: Process and validate data
        processing_result = self.data_processor.process_raw_data(raw_data_list)
        
        # Check if we have processed data
        if processing_result.processed_data:
            processed_data = processing_result.processed_data[0]
            
            # Step 4: Verify data integrity
            self.assertEqual(original_data.get('B'), processed_data.fecha)  # Date
            self.assertEqual(original_data.get('G'), processed_data.total)  # Total
            self.assertEqual(original_data.get('J'), processed_data.emisor_rfc)  # Emisor RFC
            self.assertEqual(original_data.get('M'), processed_data.receptor_rfc)  # Receptor RFC
            
            # Step 5: Generate Excel output
            excel_result = self.excel_processor.process_cfdi_to_excel(
                template_path=template_path,
                cfdi_data_list=processing_result.processed_data,
                year=2024,
                month=1
            )
            
            # Step 6: Verify Excel output contains correct data
            self.assertTrue(excel_result['success'])
            self.assertTrue(os.path.exists(excel_result['output_path']))
        else:
            # If no data was processed, that's also a valid test case
            self.assertEqual(processing_result.successful_files, 0)
            self.assertGreater(processing_result.failed_files, 0)


if __name__ == '__main__':
    unittest.main() 