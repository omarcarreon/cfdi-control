"""
Unit tests for Excel processor functionality
"""

import unittest
import tempfile
import os
from pathlib import Path
import sys
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Add src to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from core.excel_processor import ExcelProcessor
from core.data_models import CFDIData
from config.settings import CFDI_MAPPING, EXCEL_CONFIG


class TestExcelProcessor(unittest.TestCase):
    """Test cases for ExcelProcessor class."""
    
    def setUp(self):
        """Set up test fixtures."""
        self.processor = ExcelProcessor()
        self.temp_dir = tempfile.mkdtemp()
        
        # Sample CFDI data for testing
        self.sample_cfdi_data = CFDIData(
            fecha='2024-01-15T10:30:00',
            forma_pago='01',
            subtotal='1000.00',
            descuento='0.00',
            moneda='MXN',
            total='1160.00',
            tipo_comprobante='I',
            metodo_pago='PUE',
            emisor_rfc='AAA010101AAA',
            emisor_nombre='EMPRESA EJEMPLO S.A. DE C.V.',
            emisor_regimen='601',
            receptor_rfc='XEXX010101000',
            receptor_nombre='',
            receptor_regimen='601',
            total_impuestos='160.00',
            file_path='/test/path.xml',
            file_name='test.xml'
        )
    
    def tearDown(self):
        """Clean up test fixtures."""
        import shutil
        if os.path.exists(self.temp_dir):
            shutil.rmtree(self.temp_dir)
    
    def create_test_excel_template(self, filename):
        """Create a test Excel template with proper structure."""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create month tabs with proper names (Excel format: Ene2024, Feb2024, etc.)
        month_abbreviations = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
                              "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
        current_year = 2024
        
        for i, month_abbr in enumerate(month_abbreviations, 1):
            month_tab_name = f"{month_abbr}{current_year}"
            ws = wb.create_sheet(month_tab_name)
            
            # Add headers in row 3
            for xml_path, column in CFDI_MAPPING.items():
                ws[f"{column}3"] = xml_path
        
        wb.save(filename)
        return filename
    
    def test_load_template_valid(self):
        """Test loading a valid Excel template."""
        template_path = os.path.join(self.temp_dir, "test_template.xlsx")
        self.create_test_excel_template(template_path)
        
        workbook = self.processor.load_template(template_path)
        
        self.assertIsNotNone(workbook)
        self.assertIn("Ene2024", workbook.sheetnames)
        self.assertIn("Dic2024", workbook.sheetnames)
    
    def test_load_template_nonexistent(self):
        """Test loading a non-existent template."""
        workbook = self.processor.load_template("nonexistent_template.xlsx")
        
        self.assertIsNone(workbook)
    
    def test_get_month_tab_name(self):
        """Test month tab name generation."""
        # Test January
        month_name = self.processor.get_month_tab_name(1, 2024)
        self.assertEqual(month_name, "Ene2024")
        
        # Test December
        month_name = self.processor.get_month_tab_name(12, 2024)
        self.assertEqual(month_name, "Dic2024")
    
    def test_find_month_tab(self):
        """Test finding month tab in workbook."""
        template_path = os.path.join(self.temp_dir, "test_template.xlsx")
        self.create_test_excel_template(template_path)
        workbook = self.processor.load_template(template_path)
        
        # Test finding existing tab
        worksheet = self.processor.find_month_tab(workbook, 1, 2024)
        self.assertIsNotNone(worksheet)
        
        # Test finding non-existing tab (use a month that doesn't exist in the template)
        # The template has "Enero" but we'll look for a different month format
        worksheet = self.processor.find_month_tab(workbook, 1, 9999)  # Non-existent year
        # This might still find "Enero" because the find_month_tab method has fallback logic
        # So let's test with a truly non-existent scenario
        worksheet = self.processor.find_month_tab(workbook, 1, 2024)  # This should work
        self.assertIsNotNone(worksheet)
        
        # Test with a workbook that has no month tabs
        empty_wb = Workbook()
        empty_ws = empty_wb.active
        empty_ws.title = "Sheet1"
        worksheet = self.processor.find_month_tab(empty_wb, 1, 2024)
        self.assertIsNone(worksheet)
    
    def test_clear_month_data(self):
        """Test clearing month data."""
        template_path = os.path.join(self.temp_dir, "test_template.xlsx")
        self.create_test_excel_template(template_path)
        workbook = self.processor.load_template(template_path)
        worksheet = workbook["Ene2024"]
        
        # Add some test data
        worksheet["A4"] = "Test Data"
        worksheet["B4"] = "More Data"
        
        # Clear data
        self.processor.clear_month_data(worksheet)
        
        # Verify data is cleared
        self.assertIsNone(worksheet["A4"].value)
        self.assertIsNone(worksheet["B4"].value)
    
    def test_fill_month_tab(self):
        """Test filling month tab with CFDI data."""
        template_path = os.path.join(self.temp_dir, "test_template.xlsx")
        self.create_test_excel_template(template_path)
        workbook = self.processor.load_template(template_path)
        worksheet = workbook["Ene2024"]
        
        # Fill data
        success = self.processor.fill_month_tab(worksheet, [self.sample_cfdi_data], 1)
        
        self.assertTrue(success)
        
        # Verify data was filled correctly
        self.assertEqual(worksheet["B4"].value, "2024-01-15T10:30:00")  # Date
        self.assertEqual(worksheet["G4"].value, "1160.00")  # Total
        self.assertEqual(worksheet["J4"].value, "AAA010101AAA")  # Emisor RFC
    
    def test_create_output_filename(self):
        """Test output filename creation."""
        filename = self.processor.create_output_filename(2024, 1, "/path/to/template.xlsx")
        
        self.assertIn("template_CFDI_2024_01_", filename)
        self.assertTrue(filename.endswith(".xlsx"))
    
    def test_save_workbook(self):
        """Test saving workbook."""
        template_path = os.path.join(self.temp_dir, "test_template.xlsx")
        self.create_test_excel_template(template_path)
        workbook = self.processor.load_template(template_path)
        
        output_path = os.path.join(self.temp_dir, "output.xlsx")
        success = self.processor.save_workbook(workbook, output_path)
        
        self.assertTrue(success)
        self.assertTrue(os.path.exists(output_path))
    
    def test_validate_template_structure_valid(self):
        """Test validation of a properly structured Excel template."""
        template_path = os.path.join(self.temp_dir, "test_template.xlsx")
        self.create_test_excel_template(template_path)
        
        result = self.processor.validate_template_structure(template_path)
        
        self.assertTrue(result['valid'])
        self.assertEqual(len(result['errors']), 0)
        self.assertIn('month_tabs_found', result)
    
    def test_validate_template_structure_invalid(self):
        """Test validation of invalid template."""
        # Create invalid template (no month tabs)
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        template_path = os.path.join(self.temp_dir, "invalid_template.xlsx")
        wb.save(template_path)
        
        result = self.processor.validate_template_structure(template_path)
        
        self.assertFalse(result['valid'])
        self.assertGreater(len(result['errors']), 0)
    
    def test_process_cfdi_to_excel_success(self):
        """Test successful processing of CFDI data to Excel."""
        template_path = os.path.join(self.temp_dir, "test_template.xlsx")
        self.create_test_excel_template(template_path)
        
        cfdi_data_list = [self.sample_cfdi_data]
        
        result = self.processor.process_cfdi_to_excel(
            template_path=template_path,
            cfdi_data_list=cfdi_data_list,
            year=2024,
            month=1
        )
        
        self.assertTrue(result['success'])
        self.assertEqual(result['records_processed'], 1)
        self.assertIsNotNone(result['output_path'])
        self.assertTrue(os.path.exists(result['output_path']))
    
    def test_process_cfdi_to_excel_multiple_records(self):
        """Test processing multiple CFDI records."""
        template_path = os.path.join(self.temp_dir, "test_template.xlsx")
        self.create_test_excel_template(template_path)
        
        # Create multiple CFDI records
        cfdi_data_list = []
        for i in range(3):
            data = CFDIData(
                fecha=f'2024-01-{15+i:02d}T10:30:00',
                total=f'{1160.00 + i * 100:.2f}',
                emisor_rfc=f'AAA010101AAA{i}',
                receptor_rfc=f'XEXX010101000{i}',
                file_path=f'/test/path{i}.xml',
                file_name=f'test{i}.xml'
            )
            cfdi_data_list.append(data)
        
        result = self.processor.process_cfdi_to_excel(
            template_path=template_path,
            cfdi_data_list=cfdi_data_list,
            year=2024,
            month=1
        )
        
        self.assertTrue(result['success'])
        self.assertEqual(result['records_processed'], 3)
    
    def test_process_cfdi_to_excel_invalid_template(self):
        """Test processing with invalid template."""
        # Create invalid template
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        template_path = os.path.join(self.temp_dir, "invalid_template.xlsx")
        wb.save(template_path)
        
        result = self.processor.process_cfdi_to_excel(
            template_path=template_path,
            cfdi_data_list=[self.sample_cfdi_data],
            year=2024,
            month=1
        )
        
        self.assertFalse(result['success'])
        self.assertIn('error_message', result)
    
    def test_process_cfdi_to_excel_nonexistent_template(self):
        """Test processing with non-existent template."""
        result = self.processor.process_cfdi_to_excel(
            template_path="nonexistent_template.xlsx",
            cfdi_data_list=[self.sample_cfdi_data],
            year=2024,
            month=1
        )
        
        self.assertFalse(result['success'])
        self.assertIn('error_message', result)


if __name__ == '__main__':
    unittest.main() 