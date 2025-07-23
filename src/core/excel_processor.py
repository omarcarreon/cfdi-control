"""
Excel Processor for CFDI Control Application
Handles Excel template filling with CFDI data
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from pathlib import Path
from typing import List, Dict, Any, Optional
import logging
from datetime import datetime
import os

# Import our data models
from .data_models import CFDIData, ProcessingResult
from config.settings import EXCEL_CONFIG, PROCESSING_CONFIG

class ExcelProcessor:
    """Processor for Excel templates with CFDI data."""
    
    def __init__(self):
        """Initialize the Excel processor."""
        self.logger = logging.getLogger(__name__)
        
    def load_template(self, template_path: str) -> Optional[openpyxl.Workbook]:
        """
        Load Excel template file.
        
        Args:
            template_path: Path to Excel template file
            
        Returns:
            Workbook object or None if loading fails
        """
        try:
            workbook = openpyxl.load_workbook(template_path)
            self.logger.info(f"Template loaded successfully: {template_path}")
            return workbook
        except Exception as e:
            self.logger.error(f"Error loading template {template_path}: {e}")
            return None
    
    def get_month_tab_name(self, month: int, year: int = None) -> str:
        """
        Get the month tab name based on month number and year.
        
        Args:
            month: Month number (1-12)
            year: Year (optional, defaults to current year)
            
        Returns:
            Month tab name in Spanish with year (e.g., "Ene2025")
        """
        month_abbreviations = [
            "Ene", "Feb", "Mar", "Abr", "May", "Jun",
            "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"
        ]
        
        if year is None:
            from datetime import datetime
            year = datetime.now().year
            
        return f"{month_abbreviations[month - 1]}{year}"
    
    def find_month_tab(self, workbook: openpyxl.Workbook, month: int, year: int = None) -> Optional[openpyxl.worksheet.worksheet.Worksheet]:
        """
        Find the correct month tab in the workbook.
        
        Args:
            workbook: Excel workbook
            month: Month number (1-12)
            year: Year (optional, defaults to current year)
            
        Returns:
            Worksheet for the month or None if not found
        """
        month_tab_name = self.get_month_tab_name(month, year)
        
        # Try exact match first (e.g., "Ene2025")
        if month_tab_name in workbook.sheetnames:
            return workbook[month_tab_name]
        
        # Try with month number
        month_number_tab = f"{month:02d}"
        if month_number_tab in workbook.sheetnames:
            return workbook[month_number_tab]
        
        # Try with full month names (fallback)
        month_names = [
            "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
        ]
        full_month_name = month_names[month - 1]
        if full_month_name in workbook.sheetnames:
            return workbook[full_month_name]
        
        # Try with month name in different formats
        for sheet_name in workbook.sheetnames:
            if month_tab_name.lower() in sheet_name.lower():
                return workbook[sheet_name]
        
        self.logger.error(f"Month tab '{month_tab_name}' not found in template")
        return None
    
    def clear_month_data(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, start_row: int = None):
        """
        Clear existing data in the month tab (after headers).
        
        Args:
            worksheet: Excel worksheet
            start_row: Row to start clearing from (default: data_start_row from config)
        """
        if start_row is None:
            start_row = EXCEL_CONFIG['data_start_row']
        
        # Clear all data rows after headers
        for row in range(start_row, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.value = None
    
    def fill_month_tab(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                      cfdi_data_list: List[CFDIData], month: int) -> bool:
        """
        Fill the month tab with CFDI data.
        
        Args:
            worksheet: Excel worksheet to fill
            cfdi_data_list: List of CFDI data objects
            month: Month number for logging
            
        Returns:
            True if successful, False otherwise
        """
        try:
            # Clear existing data
            self.clear_month_data(worksheet)
            
            # Get starting row for data
            start_row = EXCEL_CONFIG['data_start_row']
            
            # Fill data row by row
            for row_idx, cfdi_data in enumerate(cfdi_data_list, start=start_row):
                excel_row = cfdi_data.to_excel_row()
                
                # Fill each column
                for column_letter, value in excel_row.items():
                    col_idx = openpyxl.utils.column_index_from_string(column_letter)
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    
                    # Apply formatting for numeric values
                    if column_letter in ['D', 'E', 'G', 'P']:  # Numeric columns
                        try:
                            if value and float(value) != 0:
                                cell.number_format = '#,##0.00'
                        except ValueError:
                            pass
            
            self.logger.info(f"Filled {len(cfdi_data_list)} records in month {month}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error filling month tab {month}: {e}")
            return False
    
    def create_output_filename(self, year: int, month: int, template_path: str) -> str:
        """
        Create output filename with timestamp.
        
        Args:
            year: Year
            month: Month
            template_path: Original template path
            
        Returns:
            Output filename
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        template_name = Path(template_path).stem
        
        # Create output filename
        output_filename = f"{template_name}_CFDI_{year}_{month:02d}_{timestamp}.xlsx"
        
        return output_filename
    
    def save_workbook(self, workbook: openpyxl.Workbook, output_path: str) -> bool:
        """
        Save workbook to output path.
        
        Args:
            workbook: Excel workbook
            output_path: Output file path
            
        Returns:
            True if successful, False otherwise
        """
        try:
            workbook.save(output_path)
            self.logger.info(f"Workbook saved successfully: {output_path}")
            return True
        except Exception as e:
            self.logger.error(f"Error saving workbook to {output_path}: {e}")
            return False
    
    def process_cfdi_to_excel(self, template_path: str, cfdi_data_list: List[CFDIData], 
                             year: int, month: int, output_dir: str = None) -> Dict[str, Any]:
        """
        Main method to process CFDI data and fill Excel template.
        
        Args:
            template_path: Path to Excel template
            cfdi_data_list: List of CFDI data objects
            year: Year for processing
            month: Month for processing
            output_dir: Output directory (default: same as template)
            
        Returns:
            Dictionary with processing results
        """
        result = {
            'success': False,
            'output_path': '',
            'error_message': '',
            'records_processed': 0
        }
        
        try:
            # Load template
            workbook = self.load_template(template_path)
            if not workbook:
                result['error_message'] = "No se pudo cargar la plantilla Excel"
                return result
            
            # Find month tab
            month_worksheet = self.find_month_tab(workbook, month, year)
            if not month_worksheet:
                result['error_message'] = f"No se encontró la pestaña del mes {month} para el año {year}"
                return result
            
            # Fill month tab
            if not self.fill_month_tab(month_worksheet, cfdi_data_list, month):
                result['error_message'] = "Error al llenar la pestaña del mes"
                return result
            
            # Create output filename
            output_filename = self.create_output_filename(year, month, template_path)
            
            # Determine output directory
            if output_dir is None:
                output_dir = str(Path(template_path).parent)
            
            output_path = os.path.join(output_dir, output_filename)
            
            # Save workbook
            if not self.save_workbook(workbook, output_path):
                result['error_message'] = "Error al guardar el archivo de salida"
                return result
            
            # Update result
            result['success'] = True
            result['output_path'] = output_path
            result['records_processed'] = len(cfdi_data_list)
            
            self.logger.info(f"Excel processing completed: {len(cfdi_data_list)} records processed")
            
        except Exception as e:
            result['error_message'] = f"Error inesperado: {str(e)}"
            self.logger.error(f"Unexpected error in Excel processing: {e}")
        
        return result
    
    def validate_template_structure(self, template_path: str) -> Dict[str, Any]:
        """
        Validate that the Excel template has the correct structure.
        
        Args:
            template_path: Path to Excel template
            
        Returns:
            Dictionary with validation results
        """
        validation_result = {
            'valid': False,
            'errors': [],
            'warnings': []
        }
        
        try:
            workbook = self.load_template(template_path)
            if not workbook:
                validation_result['errors'].append("No se pudo cargar la plantilla")
                return validation_result
            
            # Check if template has month tabs
            month_tabs_found = []
            current_year = datetime.now().year
            
            for month in range(1, 13):
                month_tab = self.find_month_tab(workbook, month, current_year)
                if month_tab:
                    month_tabs_found.append(month)
                else:
                    validation_result['warnings'].append(f"Pestaña del mes {month} no encontrada")
            
            if len(month_tabs_found) >= 1:
                validation_result['valid'] = True
                validation_result['month_tabs_found'] = month_tabs_found
            else:
                validation_result['errors'].append("No se encontraron pestañas de meses válidas")
            
        except Exception as e:
            validation_result['errors'].append(f"Error validando plantilla: {str(e)}")
        
        return validation_result 